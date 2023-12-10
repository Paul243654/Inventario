import csv
import sqlite3
import pandas as pd
import numpy as np
import sys
from PyQt5 import uic, QtWidgets
from PyQt5.QtWidgets import QApplication, QWidget, QInputDialog, QLineEdit, QFileDialog, QDialog, QPushButton, QLabel, QProgressBar, QTableWidgetItem, QMainWindow, QLCDNumber
from PyQt5.QtGui import QIcon
from PyQt5.QtWidgets import QMessageBox
from PyQt5.QtCore import *
from PyQt5.QtCore import QTimer, QTime, Qt
from PyQt5.QtGui import QIcon, QPixmap
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import Border, Side, PatternFill
import os
from os import remove
import time
from datetime import date
from datetime import datetime
from win32com import client
from INCIDENCIA import Consulta_incidencias
from PyQt5.QtGui import QMovie
from acercade import Acerca_de

############################################################################################################
name="BDSE",
version="3.0",
description="Inventario de producto en stock y resolución de incidencias de pedidos preparados",
author="Paul Nuñez Espinal",
author_email="Pauldav61@gmail.com",
url="https://reactivated.webnode.es",
license="Privativo",
packages=[  ("./formulario_inventario.ui"),
            ("./Base_dat.db"),
            ("./sony.ico"),
            ("./navidad-ani.gif"),
            ("./INCIDENCIA.py"),
            ("./INCIDENCIAS_SE.ui")
            ]
#####################################################################################################################
class Consulta_inventario(QDialog):
    def __init__(self):
        super().__init__()
        flags = Qt.WindowFlags()
        #self.setWindowFlags(Qt.WindowMinimizeButtonHint|Qt.WindowCloseButtonHint)
        self.setWindowFlags(Qt.WindowMinimizeButtonHint)

        # NUMERO 1/7
        #nombre_archivo=self.resolver_ruta("formulario_inventario.ui")
        #uic.loadUi(nombre_archivo, self)
        uic.loadUi("C:/Users/Paul/Desktop/PROYECTOS INFORMATICOS/FICHERO EXCEL_PYTHON/formulario_inventario.ui", self)

        self.setFixedSize(QSize(1180, 650))
        self.setStyleSheet("background-color: azure;")

        # NUMERO 2/7
        #nombre_imagen_ben=self.resolver_ruta("benceno.ico")
        #imagen_ben=QPixmap(nombre_imagen_ben)
        imagen_ben=self.setWindowIcon(QIcon("C:/Users/Paul/Desktop/PROYECTOS INFORMATICOS/FICHERO EXCEL_PYTHON/benceno.ico"))
        self.setWindowIcon(QIcon(imagen_ben))

        self.pushButton_iniciar.setStyleSheet("background-color : gainsboro")
        self.pb_obtenefichero.setStyleSheet("background-color : gainsboro")
        self.pb_ejecutar.setStyleSheet("background-color : gainsboro") 
        self.pb_exportar.setStyleSheet("background-color : gainsboro") 
        self.pushButton_close.setStyleSheet("background-color : gainsboro")
        self.pushButton_reiniciar.setStyleSheet("background-color : gainsboro") 
        self.pb_modif_libro.setStyleSheet("background-color : gainsboro")
        self.pushButton_obtener_ruta.setStyleSheet("background-color : gainsboro")
        self.pushButton_obtener_unidades.setStyleSheet("background-color : gainsboro")
        self.pushButton_asignar_ruta.setStyleSheet("background-color : gainsboro")
        self.pushButton_producto.setStyleSheet("background-color : gainsboro")
        self.pushButton_producto_stock.setStyleSheet("background-color : gainsboro")

        self.label_acceso.setStyleSheet("background-color: powder blue")
        self.label_usuario.setStyleSheet("background-color: powder blue")
        self.label_pwd.setStyleSheet("background-color: powder blue")
        self.pushButton_close.setStyleSheet("background-color: red;color: white; border: 1px solid black;")
        self.pushButton_acercade.setStyleSheet("background-color : azure")

        self.pb_obtenefichero.setDisabled(True)
        self.pb_ejecutar.setDisabled(True)
        self.pb_exportar.setDisabled(True)
        self.pushButton_reiniciar.setDisabled(True)
        self.pb_modif_libro.setDisabled(True)
        self.pushButton_obtener_ruta.setDisabled(True)
        self.pushButton_obtener_unidades.setDisabled(True)
        self.pushButton_asignar_ruta.setDisabled(True)
        self.pushButton_producto.setDisabled(True)
        self.pushButton_producto_stock.setDisabled(True)
        self.pushButton_incidencias.setDisabled(True)
        self.pushButton_seg_ped.setDisabled(True)
        self.pushButton_PROBA.setDisabled(True)

        self.lineEdit_usua.setStyleSheet("background-color: white")
        self.lineEdit_pwd.setStyleSheet("background-color: white")
        self.lineEdit_cod_producto.setStyleSheet("background-color: white")
        self.lineEdit_cod_producto_stock.setStyleSheet("background-color: white")

        # NUMERO 3/7
        #nombre_imagen=self.resolver_ruta("sony.ico")
        #imagen=QPixmap(nombre_imagen)
        imagen = QPixmap("C:/Users/Paul/Desktop/PROYECTOS INFORMATICOS/FICHERO EXCEL_PYTHON/sony.ico")
        self.label_sony.setPixmap(imagen)

        # NUMERO 4/7
        #nombre_gif=self.resolver_ruta("navidad-ani.gif")
        #self.giffs=QMovie(nombre_gif)
        self.giffs = QMovie("C:/Users/Paul/Desktop/PROYECTOS INFORMATICOS/FICHERO EXCEL_PYTHON/navidad-ani.gif")
        self.label_navidad.setMovie(self.giffs)
        self.giffs.start()
        #-------------------------------------------------------------------------------
        
        self.progressBar_calculo.setValue(0)
        #n=100
        self.pushButton_iniciar.clicked.connect(self.Gui_login)
        self.pb_obtenefichero.clicked.connect(self.obtener_direccion)
        self.pb_ejecutar.clicked.connect(self.pandas_dfa)
        self.pb_exportar.clicked.connect(self.exportar_fichero)
        self.pb_modif_libro.clicked.connect(self.modificaciones_lib)
        self.pushButton_obtener_ruta.clicked.connect(self.ejecutar_ruta)
        self.pushButton_obtener_unidades.clicked.connect(self.ejecutar_unidades)
        self.pushButton_asignar_ruta.clicked.connect(self.asignar_rutas)
        self.pushButton_reiniciar.clicked.connect(self.reiniciar_eje)
        self.pushButton_producto.clicked.connect(self.buscar_cod_prod)
        self.pushButton_producto_stock.clicked.connect(self.buscar_cod_prod_stock)
        self.pushButton_incidencias.clicked.connect(self.open_vent)
        self.pushButton_seg_ped.clicked.connect(self.obtener_clientes)
        self.pushButton_acercade.clicked.connect(self.open_about)
        self.pushButton_close.clicked.connect(self.salir_ejec)
        self.pushButton_PROBA.clicked.connect(self.fech_dfas)
        self.fec_ant=datetime(2024, 2, 3, 13, 15, 00, 00000) 
        self.fec_ah=datetime.now()
        ##############################################################################
        self.Variant_of=1111
        self.Unidades_servidas=0
        self.label_num_serv.setText("0")
        ###############################################################################
        timer=QTimer(self)
        timer.timeout.connect(self.tick)
        timer.start(1000)
        self.tick()
        #--------------------------------------------------------------------------------------------

        # NUMERO 5/7   
        #nombre_BaseDatos=self.resolver_ruta("Base_dat.db")
        #if self.fec_ah > self.fec_ant:
        #    remove(nombre_BaseDatos)
        self.rut_baseDatos="C:/Users/Paul/Desktop/PROYECTOS INFORMATICOS/FICHERO EXCEL_PYTHON/Base_dat.db"
        if self.fec_ah > self.fec_ant:
            remove(self.rut_baseDatos)

        # NUMERO 6/7   
        #nombre_ruta_inc=self.resolver_ruta("INCIDENCIA.py")
        #if self.fec_ah > self.fec_ant:
        #    remove(nombre_ruta_inc)
        self.rut_elim_inc="C:/Users/Paul/Desktop/PROYECTOS INFORMATICOS/FICHERO EXCEL_PYTHON/INCIDENCIA.py"
        if self.fec_ah > self.fec_ant:
            remove(self.rut_elim_inc)

        #------------------------------------------------------------------------------------------------------          

    def tick(self):
        hora=QTime.currentTime()
        hora_texto=hora.toString('hh:mm:ss')
        self.lcd_hora.display(hora_texto)

    def Gui_login(self):
        user_name=self.lineEdit_usua.text()
        user_pwd=self.lineEdit_pwd.text()
        try:
            if len(user_name) == 0 or len(user_pwd) == 0 :
                QMessageBox.about(self, "INFORMACION","No ha ingresado un usuario/paswword")

            elif user_name == "admin" and user_pwd == "admin_365":
                self.lineEdit_usua.clear()
                self.lineEdit_pwd.clear()
                QMessageBox.about(self, "Información","Asegurese de haber descargado e identificado los archivos necesarios.")
                self.pushButton_iniciar.setDisabled(True)
                self.pb_obtenefichero.setDisabled(False)
                self.pushButton_incidencias.setDisabled(False)
                self.pushButton_incidencias.setStyleSheet("background-color : SlateBlue;color: white; border: 1px solid black;")
            else:    
                QMessageBox.about(self, "INFORMACION","usuario/paswword incorrectos")
                self.lineEdit_user.clear()
                self.lineEdit_pwd.clear()
                self.pb_obtenefichero.setDisabled(True)
                self.pb_ejecutar.setDisabled(True)
                self.pb_exportar.setDisabled(True)
        except:
            QMessageBox.about(self, "INFORMACION","Error en la conexión sql")
    
    def conecBase(self):
        # NUMERO 7/7
        #nombre_conexion=self.resolver_ruta("Base_dat.db")
        #self.conexion = sqlite3.connect(nombre_conexion)
        self.conexion = sqlite3.connect("C:/Users/Paul/Desktop/PROYECTOS INFORMATICOS/FICHERO EXCEL_PYTHON/Base_dat.db")
        self.mi_cursor=self.conexion.cursor()  

    def desconecBase(self):
        self.conexion.commit()
        self.conexion.close()

    def crea_table(self, nombre_tabla):  
        self.conexion.execute(nombre_tabla)

    def borrar_datostabla_SU(self):
        sql= "DELETE FROM INVENTARIO;"
        self.mi_cursor.execute(sql)

    def mandar_crear(self):
        self.conecBase()
        self.crea_table(""" CREATE TABLE INVENTARIO (SSCC VARCHAR(50), PRODUCTO INTEGER, DESCRIPCION VARCHAR(50), F_CADUC DATE, ESTADO INTEGER, UNIDS_STO INTEGER, UNIDS_DES INTEGER,\
                            MOTIVO VARCHAR(50), ZONA VARCHAR(5), P1 INTEGER, P2 INTEGER, P3 INTEGER, LOTE VARCHAR(50), EMBALAJE VARCHAR(50) )""")
        self.desconecBase()
#####################################################################################################################################################################

    def borrar_datostabla_ruta(self):
        sql1= "DELETE FROM RUTA;"
        self.mi_cursor.execute(sql1)

    def borrar_datostabla_unidades(self):
        sql2= "DELETE FROM UNIDADES;"
        self.mi_cursor.execute(sql2)

######################################################################################################################################################################

    def resolver_ruta(self,ruta_relativa):
        if hasattr(sys, '_MEIPASS'):
            return os.path.join(sys._MEIPASS, ruta_relativa)
        return os.path.join(os.path.abspath('.'), ruta_relativa)

########################################################################################################################################################################
    def obtener_direccion(self):
        try:
            ruta_archivo = QFileDialog.getOpenFileName(self, "Buscar Archivo TOTALES...")
            self.ruta_archivo=ruta_archivo[0]
            self.pb_obtenefichero.setDisabled(True)
            self.progressBar_calculo.setValue(2)
            QMessageBox.about(self, "INFORMACION","Dirección del fichero Totales integrado")
            self.pb_ejecutar.setDisabled(False)
                  
        except:
            QMessageBox.about(self, "INFORMACION","Error en la obtención de la ruta del archivo csv")

    def extrae_inserta_datos(self, archivo, insertar):
        try:
            if self.fec_ah < self.fec_ant:
                lista_indice=[]
                with open (archivo, 'r') as w:
                    reader=csv.reader(w, delimiter=";")
                    for row in reader:
                        lista_indice.append(row)
                conteo=len(lista_indice)
                
                for rec in range(1,conteo-1):
                    a=lista_indice[rec][0]
                    b=lista_indice[rec][1]
                    c=lista_indice[rec][2]
                    d=lista_indice[rec][3]
                    e=lista_indice[rec][4]
                    f=lista_indice[rec][5]
                    g=lista_indice[rec][6]
                    h=lista_indice[rec][7]
                    i=lista_indice[rec][8]
                    j=lista_indice[rec][9]
                    k=lista_indice[rec][10]
                    l=lista_indice[rec][11]
                    m=lista_indice[rec][12]
                    n=lista_indice[rec][13]

                    self.conexion.execute(insertar, (a, b, c, d, e, f, g, h, i, j, k, l, m, n))
            else:
                QMessageBox.about(self, "INFORMACION","Error: actualizar archivos DLL")
                self.close()  
        except:
            QMessageBox.about(self, "INFORMACION","Error en la importación del archivo csv")
            self.close()
        
    def insertar_Tabla(self):
        
        self.borrar_datostabla_SU()
        self.extrae_inserta_datos(self.ruta_archivo, "INSERT INTO INVENTARIO(SSCC, PRODUCTO, DESCRIPCION, F_CADUC, ESTADO, UNIDS_STO, UNIDS_DES,\
                            MOTIVO, ZONA , P1, P2, P3, LOTE, EMBALAJE) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)")

    def pandas_dfa(self):
        try:
            self.pb_ejecutar.setDisabled(True)
            self.conecBase()
            self.insertar_Tabla()
            self.df = pd.read_sql_query("SELECT * FROM INVENTARIO", self.conexion)

            self.dfa=self.df[((self.df["ESTADO"]==10) | (self.df["ESTADO"]==40)) & (self.df["ZONA"]=="E")]
            
            self.dfb=self.dfa[
                                ((self.dfa["P1"]==1) & (self.dfa["P3"]<3)) |
                                ((self.dfa["P1"]==2) & (self.dfa["P3"]==0)) |
                                ((self.dfa["P1"]==3) & (self.dfa["P3"]==0)) |
                                ((self.dfa["P1"]==4) & (self.dfa["P3"]<2)) |
                                ((self.dfa["P1"]==5) & (self.dfa["P3"]>=0)) 
                                ]
            
            self.dfc=self.dfb[["PRODUCTO", "DESCRIPCION", "MOTIVO", "F_CADUC", "LOTE", "UNIDS_STO", "ZONA", "P1", "P2", "P3"]]
            self.dfc=self.dfc.assign(UBICACION="")
            
            for ind_ex, rows in self.dfc.iterrows():
                aa=self.dfc.at[self.df.index[ind_ex],'ZONA']
                bb=self.dfc.at[self.df.index[ind_ex],'P1']
                cc=self.dfc.at[self.df.index[ind_ex],'P2']
                dd=self.dfc.at[self.df.index[ind_ex],'P3']
                ubic=str(aa)+"-"+str(bb)+"-"+str(cc)+"-"+str(dd)
                self.dfc.at[ind_ex, 'UBICACION'] = ubic
            self.dfd = self.dfc.sort_values(['P1', 'P2'])
            self.progressBar_calculo.setValue(24)
            QMessageBox.about(self, "INFORMACION","Archivo generado")
            self.pb_exportar.setDisabled(False)     
        except:
            QMessageBox.about(self, "INFORMACION","Error en la generación de datos, actualize las bibliotecas correspondientes")
            self.close()

        remove(self.ruta_archivo)

    def exportar_fichero(self):
        fileName_save = QFileDialog.getSaveFileName(self,"QFileDialog.getSaveFileName()","","All Files (*);;excel File (*.xlsx)")
        nombre_tabla=fileName_save[0]
        
        if nombre_tabla != "":
            try:    
                self.pb_exportar.setDisabled(True)
                tabla_pivotante=pd.pivot_table(self.dfd, values='UNIDS_STO', index=['P1','P2','P3','LOTE', 'F_CADUC', 'MOTIVO', 'DESCRIPCION', 'PRODUCTO', 'UBICACION'], aggfunc=np.sum)
                writer=pd.ExcelWriter(nombre_tabla)
                tabla_pivotante.to_excel(writer,"Zona_E")
                tabla_pivotante.to_excel(writer,"Repaso")
                writer.save()
                self.desconecBase()
                self.progressBar_calculo.setValue(30)  
                QMessageBox.about(self, "INFORMACION", "Archivo exportado correctamente")
                self.pb_modif_libro.setDisabled(False)
                self.Variant_of=2222
            except:
                QMessageBox.about(self, "INFORMACION","Error en la exportación de la tabla pivotante")


    def modificaciones_lib(self):
        self.pb_modif_libro.setDisabled(True)
        self.label_wait.setText('Espere!!!!!!')
        try:
            ruta_modif = QFileDialog.getOpenFileName(self, "Buscar Archivo inventario...")
            self.ruta_archivo=ruta_modif[0]
            self.ruta_archivo_delete=ruta_modif[0]
            wb = load_workbook(self.ruta_archivo) 
            hoja2 = wb.create_sheet("Totales")
            #hoja3 = wb.create_sheet("Ruta")
            hoja4 = wb.create_sheet("Unidades")
        
            ws0=wb['Zona_E']
            ws1=wb['Repaso']
            ws0['K1']="UNIDADES"
            ws0['L1']="OBSERVACION"
            ws1['K1']="UNIDADES"
            ws1['L1']="OBSERVACION"
            ws2=wb['Totales']
            #ws3=wb['Ruta']
            ws4=wb['Unidades']
            
            multiple_cells_0 = ws0['A1':'L1']
            for row in multiple_cells_0:
                for cell in row:
                    cell.font = Font(name = 'Arial', size = 9, color='FFFFFF', bold=True)
                    fill = PatternFill(patternType="solid", start_color="0000FF")#   llenar
                    cell.fill = fill

            multiple_cells_1 = ws1['F1':'L1']
            for row in multiple_cells_1:
                for cell in row:
                    cell.font = Font(name = 'Arial', size = 9, color='FFFFFF', bold=True)
                    fill = PatternFill(patternType="solid", start_color="0000FF")#   llenar
                    cell.fill = fill

            wb.save(self.ruta_archivo)

            cab_ra=list(self.df)
            for m in range(0,14):
                ws2.cell(row=1, column=m+1, value=cab_ra[m])
            for i in range (len(self.df)):
                for j in range(0,14):
                    a=self.df.iloc[i,j] 
                    ws2.cell(row=i+2, column=j+1, value=a) 
        
            multiple_cells_2 = ws2['A1':'N1']
            for row in multiple_cells_2:
                for cell in row:
                    cell.font = Font(name = 'Arial', size = 9, color='FFFFFF', bold=True)
                    fill = PatternFill(patternType="solid", start_color="0000FF")#   llenar
                    cell.fill = fill
                    
            wb.save(self.ruta_archivo)

            max_fila=ws0.max_row
            for war in range(1,max_fila+1):
                for ham in range(10,13):
                    ws0.cell(row = war, column = ham).border = Border(right = Side(border_style='thin', color='000000'), bottom = Side(border_style='thin', color='000000'), left = Side(border_style='thin', color='000000'))
                for ham in range(10,13):    
                    ws1.cell(row = war, column = ham).border = Border(right = Side(border_style='thin', color='000000'), bottom = Side(border_style='thin', color='000000'), left = Side(border_style='thin', color='000000'))         
            wb.save(self.ruta_archivo)

            ws0.freeze_panes = "A2"
            ws1.freeze_panes = "A2"
            hoja2.freeze_panes = "A2"
            #hoja3.freeze_panes = "A2"
            hoja4.freeze_panes = "A2"

            wb.save(self.ruta_archivo)
            wb.close()
            self.progressBar_calculo.setValue(50)
            self.label_wait.setText('')
            QMessageBox.about(self, "INFORMACION","Archivo modificado correctamente")
        except:
            QMessageBox.about(self, "INFORMACION","Error en la modificación del excel")
            self.label_wait.setText('')

        self.pushButton_obtener_ruta.setDisabled(False)

        try:
            wb = load_workbook(self.ruta_archivo) 
            ws0=wb.worksheets[0]
            ws1=wb.worksheets[1]
            ws3=wb.worksheets[3]

            Lista_E=["A","B","C","D","E","F","G","H","I","J","K","L"]
            Lista_width_E=[3,3,3,15,11,8,36,10,11,10,9,18]
            for f in range(0,12):
                cola=Lista_E[f]
                width_cola=Lista_width_E[f]
                ws0.column_dimensions[cola].width = width_cola

            Lista_Repaso=["F","G","H","I","J","K","L"]
            Lista_width_Rep=[7,34,10,11,10,9,13]
            for g in range(0,7):
                colb=Lista_Repaso[g]
                width_colb=Lista_width_Rep[g]
                ws1.column_dimensions[colb].width = width_colb

            Lista_Unidades=["A","B","C","D","E","F","G","H","I"]
            Lista_width_uni=[14,13,11,13,40,11,11,9,11]
            for h in range(0,9):
                colc=Lista_Unidades[h]
                width_colc=Lista_width_uni[h]
                ws3.column_dimensions[colc].width = width_colc

            wb.save(self.ruta_archivo)
            wb.close()
        except:
            print("Celdas sin autofit")

    def obtener_direccion_ruta(self):
        try:
            ruta_archivo_ruta = QFileDialog.getOpenFileName(self, "Buscar Archivo RUTA...")
            self.ruta=ruta_archivo_ruta[0]
            self.pushButton_obtener_ruta.setDisabled(True) 
            QMessageBox.about(self, "INFORMACION","Dirección del fichero Ruta integrado")           
        except:
            QMessageBox.about(self, "INFORMACION","Error en la obtención de la ruta del archivo csv")


    def extrae_inserta_datos_ruta(self, archivo, insertar):
        try:
            if self.fec_ah < self.fec_ant:
                lista_indice_ruta=[]
                with open (archivo, 'r') as w:
                    reader=csv.reader(w, delimiter=";")
                    for row in reader:
                        lista_indice_ruta.append(row)
                conteo=len(lista_indice_ruta)
                
                for rec in range(1,conteo-1):
                    a=lista_indice_ruta[rec][0]
                    b=lista_indice_ruta[rec][1]
                    c=lista_indice_ruta[rec][2]
                    d=lista_indice_ruta[rec][3]
                    e=lista_indice_ruta[rec][4]
                    f=lista_indice_ruta[rec][5]
                    g=lista_indice_ruta[rec][6]
                    h=lista_indice_ruta[rec][7]
                    i=lista_indice_ruta[rec][8]
                    j=lista_indice_ruta[rec][9]
                    k=lista_indice_ruta[rec][10]
                    l=lista_indice_ruta[rec][11]
                    m=lista_indice_ruta[rec][12]
                    n=lista_indice_ruta[rec][13]
                    o=lista_indice_ruta[rec][14]

                    self.conexion.execute(insertar, (a, b, c, d, e, f, g, h, i, j, k, l, m, n, o))
            else:
              QMessageBox.about(self, "INFORMACION","Error: actualizar archivos DLL")  
        except:
            QMessageBox.about(self, "INFORMACION","Error en la importación del archivo tipo csv")

        remove(self.ruta)
        
    def insertar_Tabla_ruta(self):       
        self.borrar_datostabla_ruta()
        self.extrae_inserta_datos_ruta(self.ruta, "INSERT INTO RUTA( RUTA, FECHA, REFLIV, ORDLIV, TVACCE, CODCLI, NOMBRE, AD1CLI, VILCLI,\
                            PESO, NOMFOU, CODFOU, CODMOP, CULCOL, CPICLI) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)")

    def obtener_direccion_unidades(self):
        try:
            ruta_archivo_unidades = QFileDialog.getOpenFileName(self, "Buscar Archivo UNIDADES...")
            self.unidades=ruta_archivo_unidades[0]
            self.pushButton_obtener_unidades.setDisabled(True)
            QMessageBox.about(self, "INFORMACION","Dirección del fichero Unidades integrado")  
        except:
            QMessageBox.about(self, "INFORMACION","Error en la obtención de la ruta del archivo csv")

    def extrae_inserta_datos_unidades(self, archivo, insertar):
        try:
            if self.fec_ah < self.fec_ant:
                lista_indice_unidades=[]
                with open (archivo, 'r') as w:
                    reader=csv.reader(w, delimiter=";")
                    for row in reader:
                        lista_indice_unidades.append(row)
                conteo=len(lista_indice_unidades)
                
                for rec in range(1,conteo-1):
                    a=lista_indice_unidades[rec][0]
                    b=lista_indice_unidades[rec][1]
                    c=lista_indice_unidades[rec][2]
                    d=lista_indice_unidades[rec][3]
                    e=lista_indice_unidades[rec][4]
                    f=lista_indice_unidades[rec][5]
                    g=lista_indice_unidades[rec][6]
                    h=lista_indice_unidades[rec][7]
                    i=lista_indice_unidades[rec][8]
                    j=lista_indice_unidades[rec][9]
                    self.Unidades_servidas=self.Unidades_servidas+int(j)
                    k=lista_indice_unidades[rec][10]
                    l=lista_indice_unidades[rec][11]
                    m=lista_indice_unidades[rec][12]
                    n=lista_indice_unidades[rec][13]

                    self.conexion.execute(insertar, (a, b, c, d, e, f, g, h, i, j, k, l, m, n))
                self.label_num_serv.setText(str(self.Unidades_servidas))
            else:
              QMessageBox.about(self, "INFORMACION","Error: actualizar archivos DLL")  
        except:
            QMessageBox.about(self, "INFORMACION","Error en la importación del archivo tipo csv")

        remove(self.unidades)
        
    def insertar_Tabla_unidades(self):  
        self.borrar_datostabla_unidades()
        self.extrae_inserta_datos_unidades(self.unidades, "INSERT INTO UNIDADES(PEDIDO, REFERENCIA, ACTIVIDAD, CIRPIC, F_ENTREGA, PRODUCTO, DESCRIPCION,\
                            PCBPRO, U_PEDIDA, U_SERVIDA, ENTORNO, CAJAS, U_SUELTAS, POPCAC) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)")

    def ejecutar_ruta(self):
        self.conecBase()
        self.obtener_direccion_ruta()
        self.insertar_Tabla_ruta()
        self.desconecBase()
        self.progressBar_calculo.setValue(55)
        self.pushButton_obtener_unidades.setDisabled(False)
        
    def ejecutar_unidades(self):
        self.conecBase()
        self.obtener_direccion_unidades()
        self.insertar_Tabla_unidades()
        self.desconecBase()
        wb = load_workbook(self.ruta_archivo) 
        ws4=wb['Unidades']
        multiple_cells_4 = ws4['A1':'I1']
        for row in multiple_cells_4:
            for cell in row:
                cell.font = Font(name = 'Arial', size = 9, color='FFFFFF', bold=True)
                fill = PatternFill(patternType="solid", start_color="0000FF")#   llenar
                cell.fill = fill
        wb.save(self.ruta_archivo)
        wb.close()
        self.progressBar_calculo.setValue(60)
        self.pushButton_asignar_ruta.setDisabled(False)

    def asignar_rutas(self):
        self.pushButton_close.setDisabled(True)
        self.pushButton_asignar_ruta.setDisabled(True)
        try:
            self.label_wait.setText('Espere!!!!!!!!')
            self.conecBase()
            self.df_ruta = pd.read_sql_query("SELECT * FROM RUTA", self.conexion)
            self.df_cod_ruta=self.df_ruta[["REFLIV", "RUTA", "CODCLI", "NOMBRE"]]
            self.df_unidades = pd.read_sql_query("SELECT * FROM UNIDADES", self.conexion)
            self.df_codigo_unidades=self.df_unidades[["REFERENCIA", "F_ENTREGA", "PEDIDO", "PRODUCTO", "DESCRIPCION", "U_PEDIDA", "U_SERVIDA"]]
            self.df_cod_unidades=self.df_codigo_unidades.assign(RUTA="", CLIENTE="")
            
            rows_ruta = len(self.df_cod_ruta.axes[0]) 
            rows_unidades = len(self.df_cod_unidades.axes[0]) 
            
            for k in range(0, rows_unidades):
                value_unidades = self.df_cod_unidades.loc[k, "REFERENCIA"]
                for f in range(0, rows_ruta):
                    value_ruta= self.df_cod_ruta.loc[f, "REFLIV"]
                    if value_unidades == value_ruta:
                        elemento_ruta=(self.df_cod_ruta.loc[f, "RUTA" ]) 
                        elemento_cliente=(self.df_cod_ruta.loc[f, "CODCLI" ]) 
                        self.df_cod_unidades.at[k, 'RUTA'] = elemento_ruta
                        self.df_cod_unidades.at[k, 'CLIENTE'] = elemento_cliente
                        break
                    
            wb = load_workbook(self.ruta_archivo) 
            ws4=wb['Unidades'] 
        
            cab_tra=list(self.df_cod_unidades)
            for m in range(0,9):
                ws4.cell(row=1, column=m+1, value=cab_tra[m])
            for i in range (len(self.df_cod_unidades)):
                self.progressBar_calculo.setValue(i)
                for j in range(0,9):
                    a=self.df_cod_unidades.iloc[i,j] 
                    ws4.cell(row=i+2, column=j+1, value=a)          
            wb.save(self.ruta_archivo)

            self.desconecBase()
            self.label_wait.setText('')
            self.pushButton_reiniciar.setDisabled(False)    
            QMessageBox.about(self, "INFORMACION","Rutas asignadas") 
            
        except:
            self.label_wait.setText('')
            QMessageBox.about(self, "INFORMACION","Error en la asignación de rutas.")

        self.pushButton_close.setDisabled(False)
        self.pushButton_producto.setDisabled(False)
        self.pushButton_producto_stock.setDisabled(False)
        self.pushButton_seg_ped.setDisabled(False)


    def buscar_cod_prod(self):
        try:
            self.tableWidget_producto.clearContents()
            CODIGO_PRODUCTO=int(self.lineEdit_cod_producto.text())
            self.df_mask=self.df_cod_unidades["PRODUCTO"]==CODIGO_PRODUCTO
            self.filtered_df = self.df_cod_unidades[self.df_mask]
            
            lista_producto=self.filtered_df.to_numpy().tolist()
            w=len(lista_producto)
            
            for i in range (0, w):
                self.tableWidget_producto.insertRow(i)
                self.tableWidget_producto.setItem(i, 0, QTableWidgetItem(str(lista_producto[i][0])))
                self.tableWidget_producto.setItem(i, 1, QTableWidgetItem(str(lista_producto[i][1])))
                self.tableWidget_producto.setItem(i, 2, QTableWidgetItem(str(lista_producto[i][2])))
                self.tableWidget_producto.setItem(i, 3, QTableWidgetItem(str(lista_producto[i][3])))
                self.tableWidget_producto.setItem(i, 4, QTableWidgetItem(str(lista_producto[i][4])))
                self.tableWidget_producto.setItem(i, 5, QTableWidgetItem(str(lista_producto[i][5])))
                self.tableWidget_producto.setItem(i, 6, QTableWidgetItem(str(lista_producto[i][6])))
                self.tableWidget_producto.setItem(i, 7, QTableWidgetItem(str(lista_producto[i][7])))
                self.tableWidget_producto.setItem(i, 8, QTableWidgetItem(str(lista_producto[i][8])))         
        except:
            QMessageBox.about(self, "INFORMACION","No se ha servido el producto indicado.")


    def buscar_cod_prod_stock(self):
        try:
            self.df_stock=self.df[["PRODUCTO", "F_CADUC", "DESCRIPCION", "ESTADO", "UNIDS_STO", "ZONA", "P1", "P2", "P3", "LOTE" ]]
            self.tableWidget_producto_stock.clearContents()
            CODIGO_PRODUCTO_STOCK=int(self.lineEdit_cod_producto_stock.text())
            self.df_mask_stock=self.df_stock["PRODUCTO"]==CODIGO_PRODUCTO_STOCK
            self.filtered_df_stock = self.df_stock[self.df_mask_stock]
            
            lista_producto_stock=self.filtered_df_stock.to_numpy().tolist()
            w=len(lista_producto_stock)
            
            for i in range (0, w):
                self.tableWidget_producto_stock.insertRow(i)
                self.tableWidget_producto_stock.setItem(i, 0, QTableWidgetItem(str(lista_producto_stock[i][0])))
                self.tableWidget_producto_stock.setItem(i, 1, QTableWidgetItem(str(lista_producto_stock[i][1])))
                self.tableWidget_producto_stock.setItem(i, 2, QTableWidgetItem(str(lista_producto_stock[i][2])))
                self.tableWidget_producto_stock.setItem(i, 3, QTableWidgetItem(str(lista_producto_stock[i][3])))
                self.tableWidget_producto_stock.setItem(i, 4, QTableWidgetItem(str(lista_producto_stock[i][4])))
                self.tableWidget_producto_stock.setItem(i, 5, QTableWidgetItem(str(lista_producto_stock[i][5])))
                self.tableWidget_producto_stock.setItem(i, 6, QTableWidgetItem(str(lista_producto_stock[i][6])))
                self.tableWidget_producto_stock.setItem(i, 7, QTableWidgetItem(str(lista_producto_stock[i][7])))
                self.tableWidget_producto_stock.setItem(i, 8, QTableWidgetItem(str(lista_producto_stock[i][8])))
                self.tableWidget_producto_stock.setItem(i, 9, QTableWidgetItem(str(lista_producto_stock[i][9])))          
        except:
            QMessageBox.about(self, "INFORMACION","No figura en stock el producto indicado.")

    def obtener_clientes(self):

        desc_cl = QMessageBox.question(self, 'PyQt5 message', "¿Ha decargado el archivo CLIENTE?, ¿ha cerrado el archivo inventario?", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        self.label_wait.setText('Ejecuntando!!!!!!')
        if desc_cl == QMessageBox.Yes:
            try:
                ruta_clientes = QFileDialog.getOpenFileName(self, "Buscar Archivo CLIENTES...")
                self.ruta_clientes=ruta_clientes[0]
                QMessageBox.about(self, "INFORMACION","Dirección del fichero Clientes integrado")    
            except:
                QMessageBox.about(self, "INFORMACION","Error en la obtención de la ruta del archivo csv")
            try:   
                lista_cl=[]
                with open(self.ruta_clientes, 'r') as File:  
                    reader = csv.reader(File, delimiter=';')
                    next(reader, None)
                    for row in reader:
                        lista_cl.append(row)
                tam=len(lista_cl)
                wb = load_workbook(self.ruta_archivo) 
                ws4=wb.worksheets[3]
                row_count = ws4.max_row
                for z in range(2,row_count+1):
                    celda_ref=ws4.cell(z,1).value
                    for i in range (0,tam-1):
                        try:
                            if celda_ref== int(lista_cl[i][2]):
                                ws4.cell(z,8).value=lista_cl[i][0]
                                ws4.cell(z,9).value=int(lista_cl[i][6])
                        except:
                            pass
                wb.save(self.ruta_archivo)
                wb.close()
                QMessageBox.about(self, "INFORMACION","Campos insertados correctamente")
            except:
                QMessageBox.about(self, "INFORMACION","UnicodeDecodeError, error en la inserción de campos")
            
            remove(self.ruta_clientes)
            self.pushButton_seg_ped.setDisabled(True)
            self.label_wait.setText('')

    def fech_dfas(self):
        try:
            contador_dias=10        
            today = date.today()        
            self.dfcad=self.dfc.assign(DIAS_CAD="")
            
            for ind_ex, rows in self.dfcad.iterrows():
                
                FECHA_CAD = self.dfcad.loc[ind_ex]['F_CADUC']
                FECHA_CAD=int(FECHA_CAD)
                FECHA_CADUCIDAD=str(FECHA_CAD)

                Agno=int(FECHA_CADUCIDAD[0:4])
                Mes=int(FECHA_CADUCIDAD[4:6])
                Dia=int(FECHA_CADUCIDAD[6:8])
                
                D_CAD= (date(Agno,Mes,Dia) - today).days
                self.dfcad.at[ind_ex, 'DIAS_CAD'] = D_CAD
                self.dfcad.sort_values(['P1', 'P2'])
            df_control_Fec=(self.dfcad[(self.dfcad['DIAS_CAD']>0) & (self.dfcad['DIAS_CAD']<contador_dias)])
            
            fileNamedf_save = QFileDialog.getSaveFileName(self,"QFileDialog.getSaveFileName()","","All Files (*);;excel File (*.xlsx)")
            nombre_tabladf=fileNamedf_save[0]
            df_control_Fec.to_excel(nombre_tabladf, index=False)
        except:
            QMessageBox.about(self, "INFORMACION","Error, no se ha podido realizar el filtrado de fechas")
            

    def reiniciar_eje(self):
        try:
            remove(self.ruta_archivo_delete) 
            ret = QMessageBox.warning(self, 'REINICIO', "Va a reiniciar el programa:                                                                La información no estará disponible, asegurese de haber impreso la información necesaria y haber cerrado el archivo generado.                                                                                        Si tiene dudas, haga clic en NO y termine los procesos correspondientes.  ", QMessageBox.Yes | QMessageBox.No )
            if ret == QMessageBox.Yes:
                self.progressBar_calculo.setValue(0)
                self.pushButton_reiniciar.setDisabled(True)
                self.pushButton_producto.setDisabled(True)
                self.pushButton_producto_stock.setDisabled(True)
                self.pushButton_seg_ped.setDisabled(True)
                self.pushButton_iniciar.setDisabled(False)
                self.tableWidget_producto.clearContents()
                self.tableWidget_producto_stock.clearContents()
                self.lineEdit_cod_producto.setText("")
                self.lineEdit_cod_producto_stock.setText("")
                self.Variant_of=1111
                self.Unidades_servidas=0
                self.label_num_serv.setText("0")
                QMessageBox.about(self, "INFORMACION","Introduzca usuario y contraseña.")                    
        except:
            QMessageBox.about(self, "INFORMACION","Cierre el archivo excel generado para poder reiniciar el programa.")

    def salir_ejec(self):
        try:
            ret = QMessageBox.warning(self, 'SALIR', "Va a cerrar la aplicación:                                                                La información no estará disponible, asegurese de haber impreso la información necesaria y haber cerrado el archivo generado.                                                                                        Si tiene dudas, haga clic en NO y termine los procesos correspondientes.  ", QMessageBox.Yes | QMessageBox.No )
            if ret == QMessageBox.Yes:
                remove(self.ruta_archivo_delete)
                self.close()
        except:
            if self.Variant_of==1111:
                self.close()
            else:
                QMessageBox.about(self, "INFORMACION","Cierre el archivo excel generado para poder salir del programa.")

    def open_vent(self):
        self.window_inc = Consulta_incidencias()
        self.window_inc.show()

    def open_about(self):
        self.window_about = Acerca_de()
        self.window_about.show()

if __name__ == "__main__":
    app=QApplication(sys.argv)
    QApplication.setAttribute(Qt.AA_DisableWindowContextHelpButton)  
    GUIa=Consulta_inventario()
    GUIa.show()
    sys.exit(app.exec_())




    
        
  








